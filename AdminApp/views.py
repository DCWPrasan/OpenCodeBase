from django.db import transaction
from .serializers import (
    PerishableStocksSerializer,
    ProductStocksSerializer,
    ProductsSerializer,
    ProductsDetailsSerializer,
    RacksSerializer,
    ReportInventoryExceptionSerializer,
    SearchRacksSerializer,
    SourceSerializer,
    StocksSerializer,
    StocksHistorySerializer,
    BarcodeSerializer,
    RacksProductDetailsSerializer,
    EmployeeSerializer,
    UserSerializer,
    DesignationSerializer,
    UnitSerializer,
    ProductCategorySerializer,
    SearchEmployeeSerializer,
    ReportMaterialSerializer,
    ReportMaterialTransacrionSerializer,
    SearchMaterialSerializer,
    ReportFastReceiveMovingItemSerializer,
    CrirticalReorderItemSerializer,
    ReportEmployeeSerializer,
    ReportMaterialUsageSerializer,
    ReportDeadStockItemSerializer,
    StocksRecommendedSerializer,
)
from rest_framework.views import APIView
from rest_framework.response import Response
from .pagination import CustomPagination
from .models import (
    Products,
    Racks,
    Stocks,
    StocksHistory,
    Barcodes,
    ProductCategory,
    Employees,
    Unit,
    Source,
    BorrowedStock,
)
from AuthApp.models import Users, Designation, PERMISSION_KEYS
from AuthApp.customAuth import allowed_permission
from django.db.models import Q, Sum, Count, F, ExpressionWrapper, FloatField, Max
from django.db.models.functions import Coalesce
from django.conf import settings
from AdminApp.utils import Syserror, validate_quantity, sendEmail_template
from datetime import datetime, timedelta, date
import calendar
import re
from dateutil.relativedelta import relativedelta
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import io
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
import joblib
from AdminApp.helper import send_daily_stock_report_mail

def get_number_of_days(year, month):
    _, num_days = calendar.monthrange(year, month)
    return num_days


class DashboardView(APIView):
    def get(self, request):
        try:
            today_date = datetime.now().date()
            stock_history = (
                StocksHistory.objects.select_related("stock")
                .filter(created_at__date=today_date)
                .aggregate(
                    stock_in=Sum("quantity", filter=Q(is_stock_out=False)),
                    stock_out=Sum("quantity", filter=Q(is_stock_out=True)),
                )
            )

            prod = Products.objects.filter(net_quantity__gt=0).aggregate(
                under_stock=Count("id", filter=Q(net_quantity__lt=F("min_threshold"))),
                over_stock=Count("id", filter=Q(net_quantity__gt=F("max_threshold"))),
                in_stock=Count(
                    "id",
                    filter=Q(
                        net_quantity__lt=F("max_threshold"),
                        net_quantity__gt=F("min_threshold"),
                    ),
                ),
            )
            data = {
                "today_stock_in": stock_history["stock_in"] or 0,
                "today_stock_out": stock_history["stock_out"] or 0,
                "total_product": Products.objects.count(),
                "total_perishable_product": Products.objects.filter(
                    perishable_product=True
                ).count(),
                "under_stock": prod["under_stock"] or 0,
                "over_stock": prod["over_stock"] or 0,
                "in_stock": prod["in_stock"] or 0,
                "perishable_stock": self.get_perishable_stock(),
                "borrowed_lent_stock": self.get_borrowed_lent_stock(),
            }
            response = {
                "success": True,
                "message": "Dashboard Count",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def get_borrowed_lent_stock(self):
        borrowed_lent = (
            BorrowedStock.objects.filter(
                Q(borrowed_quantity__gt=0) | Q(lent_quantity__gt=0)
            )
            .values("product")
            .annotate(
                borrowed_product_count=Count(
                    "product", filter=Q(borrowed_quantity__gt=0), distinct=True
                ),
                lent_product_count=Count(
                    "product", filter=Q(lent_quantity__gt=0), distinct=True
                ),
                total_borrowed_quantity=Sum("borrowed_quantity"),
                total_lent_quantity=Sum("lent_quantity"),
            )
            .aggregate(
                borrowed_product_count=Count(
                    "product", filter=Q(borrowed_quantity__gt=0), distinct=True
                ),
                lent_product_count=Count(
                    "product", filter=Q(lent_quantity__gt=0), distinct=True
                ),
                total_borrowed_quantity=Sum("borrowed_quantity"),
                total_lent_quantity=Sum("lent_quantity"),
            )
        )

        # Structuring the result
        return {
            "borrowed": {
                "product_count": borrowed_lent["borrowed_product_count"] or 0,
                "total_quantity": borrowed_lent["total_borrowed_quantity"] or 0,
            },
            "lent": {
                "product_count": borrowed_lent["lent_product_count"] or 0,
                "total_quantity": borrowed_lent["total_lent_quantity"] or 0,
            },
        }

    def get_perishable_stock(self):
        today = datetime.now().date()
        date_ranges = [today + timedelta(days=i) for i in range(7)]
        # Fetch perishable stock count expiring each day
        expiring_stocks = (
            Stocks.objects.filter(
                quantity__gt=0,
                expired_date__range=[
                    today,
                    date_ranges[-1],
                ],  # Filter for the next 6 days
            )
            .values("expired_date")  # Group by expiry date
            .annotate(
                count=Count("quantity")
            )  # Count perishable stocks expiring that day
            .order_by("expired_date")
        )

        # Convert result into desired format
        datewise_count = {
            entry["expired_date"]: entry["count"] for entry in expiring_stocks
        }
        # Ensure all 7 days are covered
        counts = [datewise_count.get(date, 0) for date in date_ranges]
        dates = [date.strftime("%d-%b") for date in date_ranges]
        return {"stock": counts, "dates": dates}


class DashboardProductUsesView(APIView):
    def get(self, request):
        try:
            view_type = request.GET.get("view_type", "weekly")
            stock_type = request.GET.get("stock_type", "stockout")
            if view_type == "weekly":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=8)
                if stock_type in ["stockin", "stockout"]:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__is_stock_out=stock_type
                                        == "stockout",
                                        stocks__stockshistory__created_at__date__gte=to_date,
                                        stocks__stockshistory__created_at__date__lte=from_date,
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .order_by("-quantity")
                        .values("name", "id", "quantity")[:10]
                    )
                else:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__created_at__date__gte=to_date,
                                        stocks__stockshistory__created_at__date__lte=from_date,
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .filter(quantity=0)
                        .values("name", "id", "quantity")[:10]
                    )

            elif view_type == "monthly":
                this_month = datetime.now().month
                this_year = datetime.now().year
                if stock_type in ["stockin", "stockout"]:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__is_stock_out=stock_type
                                        == "stockout",
                                        stocks__stockshistory__created_at__month=this_month,
                                        stocks__stockshistory__created_at__year=this_year,
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .order_by("-quantity")
                        .values("name", "id", "quantity")[:10]
                    )
                else:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__created_at__month=this_month,
                                        stocks__stockshistory__created_at__year=this_year,
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .filter(quantity=0)
                        .values("name", "id", "quantity")[:10]
                    )
            else:
                this_year = datetime.now().year
                if stock_type in ["stockin", "stockout"]:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__is_stock_out=stock_type
                                        == "stockout",
                                        stocks__stockshistory__created_at__year=this_year,
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .order_by("-quantity")
                        .values("name", "id", "quantity")[:10]
                    )
                else:
                    data = (
                        Products.objects.annotate(
                            quantity=Coalesce(
                                Sum(
                                    "stocks__stockshistory__quantity",
                                    filter=Q(
                                        stocks__stockshistory__created_at__year=this_year
                                    ),
                                ),
                                0,
                                output_field=FloatField(),
                            )
                        )
                        .filter(quantity=0)
                        .values("name", "id", "quantity")[:10]
                    )

            response = {
                "success": True,
                "message": "Dashboard Count",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class DashboardProductOverStockView(APIView):
    def get(self, request):
        try:
            stock_type = request.GET.get("stock_type", "stockunder")
            if stock_type == "stockunder":
                data = (
                    Products.objects.filter(net_quantity__lt=F("min_threshold"))
                    .annotate(quantity=F("min_threshold") - F("net_quantity"))
                    .order_by("-quantity")
                    .values("name", "id", "quantity")[:10]
                )
            else:
                data = (
                    Products.objects.filter(net_quantity__gt=F("max_threshold"))
                    .annotate(quantity=(F("max_threshold") - F("net_quantity")))
                    .order_by("-quantity")
                    .values("name", "id", "quantity")[:10]
                )

            response = {
                "success": True,
                "message": "Dashboard Overstock under Stock",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class DashboardStockOverViewView(APIView):
    def get(self, request, id=None):
        try:
            view_type = request.GET.get("view_type", "weekly")

            cumulative_stock = 0
            stock_in_list = []
            stock_out_list = []
            carry_forward_list = []
            labels = []

            if view_type == "weekly":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=6)
                prev_date = from_date - timedelta(days=7)

                prev_f = Q(created_at__date__lte=prev_date)
                if id:
                    prev_f &= Q(stock__product__id=id)

                if (
                    prev_stock := StocksHistory.objects.select_related("stock")
                    .filter(prev_f)
                    .aggregate(
                        total=Coalesce(
                            Sum("quantity", filter=Q(is_stock_out=False))
                            - Sum("quantity", filter=Q(is_stock_out=True)),
                            0,
                            output_field=FloatField(),
                        )
                    )
                ):
                    cumulative_stock = prev_stock["total"]

                date_list = [
                    to_date + timedelta(days=x)
                    for x in range((from_date - to_date).days + 1)
                ]
                for i in date_list:
                    f = Q(created_at__date=i)
                    if id:
                        f &= Q(stock__product__id=id)
                    stock_item = (
                        StocksHistory.objects.select_related("stock")
                        .filter(f)
                        .values("created_at__date")
                        .aggregate(
                            stock_in=Coalesce(
                                Sum("quantity", filter=Q(is_stock_out=False)),
                                0,
                                output_field=FloatField(),
                            ),
                            stock_out=Coalesce(
                                Sum("quantity", filter=Q(is_stock_out=True)),
                                0,
                                output_field=FloatField(),
                            ),
                        )
                    )
                    stock_in = stock_item["stock_in"]
                    stock_out = stock_item["stock_out"]
                    cumulative_stock = cumulative_stock + stock_in - stock_out

                    stock_in_list.append(stock_in)
                    stock_out_list.append(stock_out)
                    carry_forward_list.append(cumulative_stock)
                    labels.append(i.strftime("%d-%b"))

            elif view_type == "monthly":
                m = datetime.now().strftime("%m")
                y = datetime.now().strftime("%Y")
                this_year_date = datetime(int(y) - 1, int(m), 1)
                prev_year_date = this_year_date - timedelta(days=1)
                prev_f = Q(created_at__lte=prev_year_date)
                if id:
                    prev_f &= Q(stock__product__id=id)
                if (
                    prev_stock := StocksHistory.objects.select_related("stock")
                    .filter(prev_f)
                    .aggregate(
                        total=Coalesce(
                            Sum("quantity", filter=Q(is_stock_out=False))
                            - Sum("quantity", filter=Q(is_stock_out=True)),
                            0,
                            output_field=FloatField(),
                        )
                    )
                ):
                    cumulative_stock = prev_stock["total"]

                for i in range(13):
                    if i != 0:
                        this_year_date = this_year_date + timedelta(
                            days=get_number_of_days(
                                this_year_date.year, this_year_date.month
                            )
                        )
                    this_year = this_year_date.strftime("%Y")
                    this_month = this_year_date.strftime("%m")

                    lbl_mon = date(int(this_year), int(this_month), 1).strftime("%b")
                    lbl_year = date(int(this_year), int(this_month), 1).strftime("%Y")

                    labels.append(f"{lbl_mon}-{lbl_year}")
                    f = Q(created_at__month=this_month, created_at__year=this_year)
                    if id:
                        f &= Q(stock__product__id=id)

                    stock_history = (
                        StocksHistory.objects.select_related("stock")
                        .filter(f)
                        .aggregate(
                            stock_in=Coalesce(
                                Sum("quantity", filter=Q(is_stock_out=False)),
                                0,
                                output_field=FloatField(),
                            ),
                            stock_out=Coalesce(
                                Sum("quantity", filter=Q(is_stock_out=True)),
                                0,
                                output_field=FloatField(),
                            ),
                        )
                    )

                    stock_in = stock_history["stock_in"]
                    stock_out = stock_history["stock_out"]
                    cumulative_stock = cumulative_stock + stock_in - stock_out

                    stock_in_list.append(stock_in)
                    stock_out_list.append(stock_out)
                    carry_forward_list.append(cumulative_stock)
            else:
                from_year = datetime.now().strftime("%Y")
                to_year = f"{int(from_year)-10}"
                prev_year = f"{int(from_year)-11}"

                f = Q(created_at__year__gte=to_year)
                if id:
                    f &= Q(stock__product__id=id)
                stock_history = (
                    StocksHistory.objects.select_related("stock")
                    .filter(f)
                    .values("created_at__year")
                    .annotate(
                        stock_in=Coalesce(
                            Sum("quantity", filter=Q(is_stock_out=False)),
                            0,
                            output_field=FloatField(),
                        ),
                        stock_out=Coalesce(
                            Sum("quantity", filter=Q(is_stock_out=True)),
                            0,
                            output_field=FloatField(),
                        ),
                    )
                    .order_by("created_at__year")
                )

                prev_f = Q(created_at__year__lte=prev_year)
                if id:
                    prev_f &= Q(stock__product__id=id)

                if (
                    prev_stock := StocksHistory.objects.select_related("stock")
                    .filter(prev_f)
                    .aggregate(
                        total=Coalesce(
                            Sum("quantity", filter=Q(is_stock_out=False))
                            - Sum("quantity", filter=Q(is_stock_out=True)),
                            0,
                            output_field=FloatField(),
                        )
                    )
                ):
                    cumulative_stock = prev_stock["total"]

                for stock_item in stock_history:
                    stock_in = stock_item["stock_in"]
                    stock_out = stock_item["stock_out"]
                    cumulative_stock = cumulative_stock + stock_in - stock_out

                    stock_in_list.append(stock_in)
                    stock_out_list.append(stock_out)
                    carry_forward_list.append(cumulative_stock)
                    labels.append(stock_item["created_at__year"])

            data = {
                "labels": labels,
                "total_stock_in": sum(stock_in_list),
                "total_stock_out": sum(stock_out_list),
                "total_product_quantity": cumulative_stock,
                "quantity": sum(stock_out_list),
                "series": [
                    {
                        "name": "Total Stock",
                        "data": carry_forward_list,
                        "type": "column",
                    },
                    {"name": "Stock IN", "data": stock_in_list, "type": "area"},
                    {"name": "Stock OUT", "data": stock_out_list, "type": "line"},
                ],
            }
            response = {
                "success": True,
                "message": "Dashboard Count",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class RacksView(APIView, CustomPagination):
    serializer_class = RacksSerializer

    def get(self, request, id=None):
        try:
            if id:

                if rack := Racks.objects.filter(
                    Q(id=id) | Q(barcode__barcode_no=id)
                ).first():
                    products = Products.objects.filter(stocks__rack=rack).distinct()
                    product_data = RacksProductDetailsSerializer(
                        products, many=True, context={"rack_id": id}
                    ).data
                    response = {
                        "success": True,
                        "message": "Rack get successfully",
                        "data": {
                            "rack": {
                                "rack_no": rack.rack_no,
                                "barcode": rack.barcode.barcode_no,
                            },
                            "products": product_data,
                        },
                    }
                    return Response(response, status=200)
                else:
                    response = {"success": False, "message": "Rack Not found"}
                    return Response(response, status=400)

            if query := request.GET.get("query"):
                instance = Racks.objects.select_related('barcode').filter(
                    Q(rack_no__icontains=query)
                    | Q(barcode__barcode_no__icontains=query)
                )
            else:
                instance = Racks.objects.select_related('barcode')

            instance = instance.annotate(
                product=Count("stocks__product", distinct=True),
            ).order_by("rack_no")

            if export := request.GET.get("export") == "true":
                file_name = f"rack-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_rack_to_excel(instance, file_name)
            
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Rack List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        
    def export_rack_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"
            # Headers
            headers = [
                "Rack",
                "barcode",
                "Material Count"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.rack_no,
                    row.barcode.barcode_no,
                    row.product
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)

    @allowed_permission("manage_rack")
    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                rack_no = data.get("rack_no", None)
                barcode_no = data.get("barcode_no", None)
                if not all([rack_no, barcode_no]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                barcode = Barcodes.objects.filter(
                    barcode_no=barcode_no.upper(), is_product_type=False
                ).first()
                if not barcode:
                    response = {
                        "success": False,
                        "message": "Invalid barcode",
                    }
                    return Response(response, status=400)
                if Racks.objects.filter(
                    Q(rack_no=rack_no) | Q(barcode=barcode)
                ).exists():
                    response = {
                        "success": False,
                        "message": "Rack no or barcode no already exists",
                    }
                    return Response(response, status=400)
                instance = Racks.objects.create(rack_no=rack_no, barcode=barcode)
                barcode.status = "Used"
                barcode.save()
                response = {
                    "success": True,
                    "message": "Rack Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    @allowed_permission("manage_rack")
    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                rack_id = data.get("rack_id", None)
                rack_no = data.get("rack_no", None)
                barcode_no = data.get("barcode_no", None)
                if not all([rack_no, barcode_no, rack_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                instance = Racks.objects.get(id=rack_id)
                if (
                    Racks.objects.filter(rack_no=rack_no)
                    .exclude(rack_no=instance.rack_no)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Rack No already exists",
                    }
                    return Response(response, status=400)
                barcode = Barcodes.objects.filter(
                    barcode_no=barcode_no.upper(), is_product_type=False
                ).first()
                if not barcode:
                    response = {
                        "success": False,
                        "message": "Invalid barcode",
                    }
                    return Response(response, status=400)

                if (
                    Racks.objects.filter(barcode=barcode)
                    .exclude(barcode=instance.barcode)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Barcode already exists",
                    }
                    return Response(response, status=400)
                instance.barcode = barcode
                instance.rack_no = rack_no
                instance.save()
                if barcode.status != "Used":
                    barcode.status = "Used"
                    barcode.save()
                response = {
                    "success": True,
                    "message": "Rack Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def delete(self, request, id):
        try:
            with transaction.atomic():
                instance = Racks.objects.get(id=id)
                response = {
                    "success": True,
                    "message": "Rack Deleted Successfully",
                }
                instance.delete()
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ProductsView(APIView, CustomPagination):
    serializer_class = ProductsSerializer

    def get(self, request, id=None):
        try:
            if id:
                if rack_id := request.GET.get("rack_id", None):
                    stocks = (
                        Stocks.objects.select_related("product", "rack")
                        .filter(product=id, rack=rack_id, quantity__gt=0)
                        .values("barcode__barcode_no", "quantity", "rack__rack_no")
                        .order_by("-quantity")
                    )
                    response = {
                        "success": True,
                        "message": "Products Stock List Get Succesfully",
                        "data": stocks,
                    }
                    return Response(response, status=200)
                instance = Products.objects.get(id=id)
                racks = (
                    Racks.objects.filter(stocks__product__id=id)
                    .annotate(net_quantity=Sum("stocks__quantity"))
                    .filter(net_quantity__gt=0)
                    .values("id", "rack_no", "net_quantity")
                )
                history = (
                    StocksHistory.objects.select_related("stock")
                    .filter(stock__product__id=id)
                    .order_by("-created_at")[:10]
                )
                recommended = (
                    Stocks.objects.select_related("rack", "product", "barcode")
                    .filter(product=instance, quantity__gt=0)
                    .order_by("expired_date", "created_at")[:10]
                )
                product_serializer = ProductsDetailsSerializer(instance)
                history_serializer = StocksHistorySerializer(history, many=True)
                recommended_stock = StocksRecommendedSerializer(recommended, many=True)
                data = {
                    "product": product_serializer.data,
                    "racks": racks,
                    "recommended_stock": recommended_stock.data,
                    "histories": history_serializer.data,
                }
                response = {
                    "success": True,
                    "message": "Products Get Succesfully",
                    "data": data,
                }
                return Response(response, status=200)
            f = Q()
            if query := request.GET.get("query"):
                f &= Q(
                    Q(name__icontains=query)
                    | Q(net_quantity__icontains=query)
                    | Q(ucs_code__icontains=query)
                    | Q(description__icontains=query)
                    | Q(description_sap__icontains=query)
                    | Q(price__icontains=query)
                    | Q(category__name__icontains=query)
                )
            
            if stock_type := request.GET.get("stock_type"):
                if stock_type == "Under Stock":
                    f &= Q(net_quantity__lt=F("min_threshold"))
                elif stock_type == "Over Stock":
                    f &= Q(net_quantity__gt=F("max_threshold"))
                elif stock_type == "In Stock":
                    f &= Q(
                        net_quantity__gte=F("min_threshold"),
                        net_quantity__lte=F("max_threshold"),
                    )
            
            if status := request.GET.get("status"):
                if status == "Active":
                    f &= Q(is_active=True)
                elif status == "Inactive":
                    f &= Q(is_active=False)
            else:
                f &= Q(is_active=True)
            
            if perishability := request.GET.get("perishability"):
                if perishability == "Perishable Material":
                    f &= Q(perishable_product=True)
                elif perishability == "Non-Perishable Material":
                    f &= Q(perishable_product=False)
            
            if unit_type := request.GET.get("unit_type"):
                if unit_type == "Multiple Quantities":
                    f &= Q(is_mutli_type_unit=True)
                elif unit_type == "Single Quantity":
                    f &= Q(is_mutli_type_unit=False)
            
            instance = (
                Products.objects.select_related().filter(f).order_by("-created_at")
            )
            if export := request.GET.get("export") == "true":
                file_name = f"material-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_material_to_excel(instance, file_name)
            
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Products List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        
    
    def export_material_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"

            # Headers
            headers = [
                "Material",
                "UCS Code",
                "Price",
                "Net Quantity",
                "Min Threshold",
                "Max Threshold",
                "Category",
                "Unit",
                "Unit Type",
                "Perishability",
                "Status"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.name,
                    row.ucs_code,
                    row.price,
                    row.net_quantity,
                    row.min_threshold,
                    row.max_threshold,
                    row.category.name,
                    row.unit.name,
                    f"{'Multiple Quantities' if row.is_mutli_type_unit else 'Single Quantity' }",
                    f"{'Perishable Material' if row.perishable_product else 'Non-Perishable Material' }",
                    f"{'Active' if row.is_active else 'Inactive' }",
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)



    @allowed_permission("manage_material")
    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                name = data.get("name", None) or None
                category_id = data.get("category", None) or None
                description = data.get("description", None) or None
                sap_description = data.get("sap_description", None) or None
                if price := data.get("price", None) or 0:
                    price = int(price) if isinstance(price, str) else price
                min_threshold = data.get("min_threshold", None) or None
                max_threshold = data.get("max_threshold", None) or None
                unit = data.get("unit", None) or None
                ved_category = data.get("ved_category", None) or None
                perishable_product = data.get("perishable_product", False)
                if lead_time := data.get("lead_time", None) or 1:
                    lead_time = (
                        int(lead_time) if isinstance(lead_time, str) else lead_time
                    )
                ucs_code = data.get("ucs_code", None) or None
                if not all(
                    [
                        name,
                        min_threshold,
                        ucs_code,
                        unit,
                    ]
                ):
                    response = {
                        "success": False,
                        "message": "Required All Mandatory fields",
                    }
                    return Response(response, status=400)
                if price and price <= 0:
                    response = {
                        "success": False,
                        "message": "Price must be greater than zero",
                    }
                    return Response(response, status=400)

                if price and lead_time <= 0:
                    response = {
                        "success": False,
                        "message": "Lead Time must be greater than zero",
                    }
                    return Response(response, status=400)

                if ved_category and ved_category not in [
                    "Vital",
                    "Essential",
                    "Desirable",
                ]:
                    response = {"success": False, "message": "Invalid VED category"}
                    return Response(response, status=400)

                if unit_type := data.get("unit_type", "mutli"):
                    unit_type = isinstance(unit_type, str) and unit_type == "multi"
                if category_id:
                    product_category = ProductCategory.objects.get(id=category_id)
                else:
                    product_category = None
                unit = Unit.objects.filter(id=unit).first()
                if unit is None:
                    response = {"success": False, "message": "Invalid Unit"}
                    return Response(response, status=400)

                TOTAL_ALLOWED_MATERIAL = settings.TOTAL_ALLOWED_MATERIAL
                if Products.objects.all().count() >= TOTAL_ALLOWED_MATERIAL:
                    response = {
                        "success": False,
                        "message": f"You have reached the maximum limit of {settings.TOTAL_ALLOWED_MATERIAL} materials allowed. Please contact the administrator for assistance.",
                    }
                    return Response(response, status=400)

                instance = Products.objects.create(
                    name=name,
                    description=description,
                    description_sap=sap_description,
                    min_threshold=min_threshold,
                    max_threshold=max_threshold,
                    price=price,
                    lead_time=lead_time,
                    ved_category=ved_category,
                    is_mutli_type_unit=unit_type,
                    unit=unit,
                    category=product_category,
                    ucs_code=ucs_code,
                    perishable_product=perishable_product,
                    is_active=True
                )
                response = {
                    "success": True,
                    "message": "Products Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    @allowed_permission("manage_material")
    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                product_id = data.get("product_id", None)
                name = data.get("name", None) or None
                category_id = data.get("category", None) or None
                description = data.get("description", None) or None
                sap_description = data.get("sap_description", None) or None
                if price := data.get("price", None) or 0:
                    price = int(price) if isinstance(price, str) else price
                min_threshold = data.get("min_threshold", None) or None
                max_threshold = data.get("max_threshold", None) or None
                unit = data.get("unit", None) or None
                ved_category = data.get("ved_category", None) or None
                perishable_product = data.get("perishable_product", False)
                is_active = data.get("is_active", True)
                if not isinstance(is_active, bool):
                    response = {
                        "success": False,
                        "message": "Material status must be true or false",
                    }
                    return Response(response, status=400)
                
                if lead_time := data.get("lead_time", None) or 1:
                    lead_time = (
                        int(lead_time) if isinstance(lead_time, str) else lead_time
                    )
                ucs_code = data.get("ucs_code", None) or None
                if not all([product_id, name, min_threshold, ucs_code, unit]):
                    response = {
                        "success": False,
                        "message": "Required All fields",
                    }
                    return Response(response, status=400)
                if price and price <= 0:
                    response = {
                        "success": False,
                        "message": "Price must be greater than zero",
                    }
                    return Response(response, status=400)

                if lead_time and lead_time <= 0:
                    response = {
                        "success": False,
                        "message": "Lead Time must be greater than zero",
                    }
                    return Response(response, status=400)

                if ved_category not in ["Vital", "Essential", "Desirable"]:
                    response = {
                        "success": False,
                        "message": "Invalid VED category",
                    }
                    return Response(response, status=400)

                if unit_type := data.get("unit_type", False):
                    unit_type = isinstance(unit_type, str) and unit_type == "multi"
                if category_id:
                    product_category = ProductCategory.objects.get(id=category_id)
                else:
                    product_category = None
                unit = Unit.objects.filter(id=unit).first()
                instance = Products.objects.get(id=product_id)
                if (not is_active) and (instance.net_quantity > 0):
                    response = {
                        "success": False,
                        "message": f"Material have {instance.net_quantity} stocks ? Please clear all stocks then try inactive",
                    }
                    return Response(response, status=400)
                
                instance.name = name
                instance.description = description
                instance.description_sap = sap_description
                instance.min_threshold = min_threshold
                instance.max_threshold = max_threshold
                instance.price = price
                instance.is_mutli_type_unit = unit_type
                instance.unit = unit
                instance.ved_category = ved_category
                instance.lead_time = lead_time
                instance.category = product_category
                instance.ucs_code = ucs_code
                instance.perishable_product = perishable_product
                instance.is_active=is_active
                instance.save()
                response = {
                    "success": True,
                    "message": "Product Updated Successfully",
                    "data": ProductsDetailsSerializer(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def delete(self, request, id):
        try:
            with transaction.atomic():
                instance = Products.objects.get(id=id)
                response = {
                    "success": True,
                    "message": "Products Deleted Successfully",
                }
                instance.delete()
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    # upload excel file
    @allowed_permission("manage_material")
    def patch(self, request):
        try:
            with transaction.atomic():
                file = request.FILES.get("file", None)
                if not file:
                    response = {
                        "success": False,
                        "message": "Required a Excel File",
                    }
                    return Response(response, status=400)
                df = pd.read_excel(file)
                error_list = []
                for index, data in df.iterrows():
                    name = data.get("name", None)
                    category = data.get("category", None)
                    description = data.get("description", None)
                    sap_description = data.get("sap_description", None)
                    if price := data.get("price", None):
                        price = int(price) if isinstance(price, str) else price
                    min_threshold = data.get("min_threshold", None)
                    max_threshold = data.get("max_threshold", None)
                    unit = data.get("unit", None)
                    ved_category = data.get("ved_category", None)
                    if lead_time := data.get("lead_time", None):
                        lead_time = (
                            int(lead_time) if isinstance(lead_time, str) else lead_time
                        )
                    ucs_code = data.get("ucs_code", None)
                    if not all(
                        [
                            name,
                            ved_category,
                            lead_time,
                            description,
                            sap_description,
                            category,
                            price,
                            min_threshold,
                            max_threshold,
                            ucs_code,
                            unit,
                        ]
                    ):
                        errors = {"row": index + 1, "message": "Required All fields"}
                        error_list.append(errors)
                        continue
                    if price <= 0:
                        errors = {
                            "row": index + 1,
                            "message": "Price must be greater than zero",
                        }
                        error_list.append(errors)
                        continue
                    if lead_time <= 0:
                        errors = {
                            "row": index + 1,
                            "message": "Lead Time must be greater than zero",
                        }
                        error_list.append(errors)
                        continue
                    if ved_category not in ["Vital", "Essential", "Desirable"]:
                        errors = {"row": index + 1, "message": "Invalid VED category"}
                        error_list.append(errors)
                        continue
                    if Products.objects.filter(ucs_code=ucs_code).exists():
                        errors = {
                            "row": index + 1,
                            "message": "UCS Code already exists",
                        }
                        error_list.append(errors)
                        continue

                    if unit_type := data.get("unit_type", False):
                        unit_type = isinstance(unit_type, str) and unit_type == "Multi"

                    (
                        product_category,
                        created_p_cat,
                    ) = ProductCategory.objects.get_or_create(name=category)

                    instance = Products.objects.create(
                        name=name,
                        description=description,
                        description_sap=sap_description,
                        min_threshold=min_threshold,
                        max_threshold=max_threshold,
                        price=price,
                        lead_time=lead_time,
                        ved_category=ved_category,
                        is_mutli_type_unit=unit_type,
                        unit=unit,
                        category=product_category or created_p_cat,
                        ucs_code=ucs_code,
                    )
                response = {
                    "success": True,
                    "message": "Products Created  Successfully",
                    "data": error_list,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class StocksView(APIView, CustomPagination):
    serializer_class = StocksSerializer

    def get(self, request, id=None):
        try:
            if id:
                instance = Stocks.objects.get(id=id)
                serializer = ProductsDetailsSerializer(instance)
                response = {
                    "success": True,
                    "message": "Products Get Succesfully",
                    "data": serializer.data,
                }
                return Response(response, status=200)
            
            if barcode := request.GET.get("barcode", None):
                return self.validate_stock_out_barcode(request, barcode)
            
            instance = Products.objects.select_related().order_by("-created_at")
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Products List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    # stock in
    @allowed_permission("manage_stockin")
    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                user = request.user
                stockin_type = data.get("stockin_type", None)
                if not stockin_type:
                    response = {
                        "success": False,
                        "message": "Required Stockin type",
                    }
                    return Response(response, status=400)
                if stockin_type not in [
                    "NEW MATERIAL",
                    "RETURN MATERIAL",
                    "BORROWING MATERIAL",
                    "RECEIVE LENT MATERIAL",
                ]:
                    response = {
                        "success": False,
                        "message": f"Invalid Stockin type {stockin_type}",
                    }
                    return Response(response, status=400)
                if stock_list := data.get("stock_list", []):
                    for i in stock_list:
                        result = self.create_stock_in(i, user, stockin_type)
                        if isinstance(result, Response):
                            return result
                    response = {
                        "success": True,
                        "message": "Stock IN  Successfully",
                        "data": [],
                    }
                    return Response(response, status=200)
                else:
                    return self.create_stock_in(data, user, stockin_type, True)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    # STOCK OUT SINGLE RACK
    def put(self, request):
        try:
            if request.user.role == "User":
                has_permissions = request.user.user_permission.get(
                    "manage_stockout", False
                ) or request.user.user_permission.get("consumption_stockout", False)
                if not has_permissions:
                    return Response(
                        {
                            "success": False,
                            "message": "You do not have permission to perform this action.",
                        },
                        status=400,
                    )
            with transaction.atomic():
                data = request.data
                stock_list = data.get("stock_list", [])
                stockout_type = data.get("stockout_type", None)
                if not stockout_type:
                    response = {
                        "success": False,
                        "message": "Required Stockout type",
                    }
                    return Response(response, status=400)
                if stockout_type not in [
                    "MATERIAL CONSUMPTION",
                    "RETURN BORROWED MATERIAL",
                    "LENDING MATERIAL",
                ]:
                    response = {
                        "success": False,
                        "message": f"Invalid Stockout type {stockout_type}",
                    }
                    return Response(response, status=400)
                if (
                    request.user.role == "User"
                    and request.user.user_permission.get("consumption_stockout", False)
                    and stockout_type != "MATERIAL CONSUMPTION"
                ):
                    response = {
                        "success": False,
                        "message": f"You have only permission for Stockout type MATERIAL CONSUMPTION",
                    }
                    return Response(response, status=400)
                if not stock_list:
                    response = {
                        "success": False,
                        "message": "Required Stock OUT List",
                    }
                    return Response(response, status=400)

                stock_ids = [
                    s["stock_id"] for s in stock_list if s["select_quantity"] > 0
                ]
                stocks = (
                    Stocks.objects.select_related("product", "rack")
                    .filter(id__in=stock_ids)
                    .order_by("quantity")
                )
                lowStockProduct = []
                obsolete_inventory_stock = []
                self_employee = None
                if request.user.role == "User" and request.user.user_permission.get(
                    "consumption_stockout", False
                ):
                    self_employee, created = Employees.objects.get_or_create(
                        personnel_number=request.user.personnel_number,
                        defaults={
                            "name": request.user.name,
                            "phone": request.user.mobile_number,
                        },
                    )
                for i in stocks:
                    quantity = self.get_value_from_stock_list(
                        stock_list, i.id, "select_quantity"
                    )
                    employee_id = self.get_value_from_stock_list(
                        stock_list, i.id, "employee_id"
                    )
                    product = Products.objects.get(id=i.product.id)
                    if self_employee is None:
                        employee = Employees.objects.filter(id=employee_id).first()
                        if employee is None:
                            raise ValueError("Invalid Employee ID")
                    else:
                        employee = self_employee

                    if stockout_type == "MATERIAL CONSUMPTION":
                        purpose = self.get_value_from_stock_list(
                            stock_list, i.id, "purpose"
                        )
                        source = None
                    else:
                        source_id = self.get_value_from_stock_list(
                            stock_list, i.id, "source"
                        )
                        source = Source.objects.filter(
                            id=source_id, is_central_store=False
                        ).first()
                        if not source:
                            raise ValueError("Invalid source")

                        if stockout_type == "LENDING MATERIAL":
                            lent_stock, created = BorrowedStock.objects.get_or_create(
                                product=product, source=source
                            )
                            lent_stock.lent_quantity += quantity
                            lent_stock.save()
                            purpose = f"Lent to {source.name}"

                        else:  # Returning borrowed material
                            purpose = f"Returned to {source.name}"
                            borrowed_stock = BorrowedStock.objects.filter(
                                product=product, source=source
                            ).first()

                            if not borrowed_stock:
                                raise ValueError(
                                    f"There is no borrowed material from source {source.name} for {product.name}"
                                )

                            if quantity > borrowed_stock.borrowed_quantity:
                                raise ValueError(
                                    f"Returning more quantity than borrowed for {product.name}. "
                                    f"Actual borrowed quantity: {borrowed_stock.borrowed_quantity}"
                                )

                            borrowed_stock.borrowed_quantity -= quantity
                            borrowed_stock.save()
    
                    if 0 < quantity <= i.quantity:
                        i.quantity -= quantity
                        i.save()
                        product.net_quantity -= quantity
                        product.save()
                        obsolete_inventory_barcode = None
                        if product.perishable_product:
                            if (
                                lastStock := Stocks.objects.filter(
                                    product=i.product,
                                    quantity__gt=0,
                                    expired_date__isnull=False,
                                )
                                .exclude(id__in=obsolete_inventory_stock)
                                .order_by("expired_date")
                                .first()
                            ):
                                if lastStock != i:
                                    obsolete_inventory_barcode = (
                                        lastStock.barcode.barcode_no
                                    )
                                    print(
                                        "obsolete_inevettory ",
                                        obsolete_inventory_barcode,
                                    )
                            obsolete_inventory_stock.append(i.id)

                        StocksHistory.objects.create(
                            stock=i,
                            source=source,
                            quantity=quantity,
                            history_type=stockout_type,
                            product_quantity=product.net_quantity,
                            user=request.user,
                            is_stock_out=True,
                            employee=employee,
                            purpose=purpose,
                            obsolete_inventory_barcode=obsolete_inventory_barcode,
                        )

                        if (
                            product.net_quantity < product.min_threshold
                            and product not in lowStockProduct
                        ):
                            lowStockProduct.append(product)

                if len(lowStockProduct) > 0:
                    self.send_low_stock_email_alert(lowStockProduct)
                response = {
                    "success": True,
                    "message": "Stock OUT successful",
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def get_value_from_stock_list(
        self, stock_list: list, stock_id: str, required_value
    ):
        default_output = None
        if required_value == "select_quantity":
            default_output = 0
        return next(
            (s.get(required_value) for s in stock_list if s["stock_id"] == stock_id),
            default_output,
        )

    def create_stock_in(
        self, data, user, stockin_type="NEW MATERIAL", is_single_stock=False
    ):
        try:
            required_fields = ["product_id", "rack_id", "barcode_no", "source"]
            if not all(data.get(field) for field in required_fields):
                raise ValueError("All required fields must be provided.")

            quantity = validate_quantity(str(data.get("quantity", 1)))
            if not quantity or quantity <= 0:
                raise ValueError("Quantity must be greater than zero.")

            rack = Racks.objects.get(id=data["rack_id"])
            product = Products.objects.get(id=data["product_id"])
            barcode = Barcodes.objects.filter(
                barcode_no=data["barcode_no"].upper()
            ).first()

            if not barcode or not barcode.is_product_type:
                raise ValueError("Invalid or non-material barcode.")
        
            source = Source.objects.filter(
                    id=data["source"],
                    is_central_store=stockin_type not in ["BORROWING MATERIAL", "RECEIVE LENT MATERIAL"],
                ).first()
            if not source:
                raise ValueError("Source not found.")

            expired_date = None
            lent_material_stock = None
            is_used_stock_barcode = False # check stock in for existing used barcode
            if stockin_type == "RETURN MATERIAL":
                stock = barcode.stocks.first()
                if not stock:
                    return Response(
                        {
                            "success": False,
                            "message": f"Stock not found with {barcode.barcode_no} barcode",
                        },
                        status=400,
                    )

                if not stock.stockshistory_set.filter(
                    history_type="MATERIAL CONSUMPTION", is_stock_out=True
                ).exists():
                    return Response(
                        {
                            "success": False,
                            "message": f"Stock history not found with {barcode.barcode_no} barcode for MATERIAL CONSUMPTION",
                        },
                        status=400,
                    )

            else:
                if product.is_mutli_type_unit:
                    if barcode.status == "Used":
                        if product.perishable_product:
                            response = {
                                "success": False,
                                "message": "Barcode already used, Can not accept duplicate Barcode for perishable material",
                            }
                            return Response(response, status=400)
                
                        is_same_product = barcode.stocks.all().exclude(
                            product__id=product.id
                        )
                        if is_same_product.count() != 0:
                            response = {
                                "success": False,
                                "message": "Barcode already used in another material.",
                            }
                            return Response(response, status=400)
                        else:
                            is_used_stock_barcode = True
                else:
                    if Stocks.objects.filter(barcode=barcode).exists():
                        raise ValueError("Barcode already used.")

                if product.perishable_product:
                    expired_date = data.get("expired_date", None)
                    if not expired_date:
                        raise ValueError(
                            "Perishable material can't be added directly without expired date."
                        )
                    try:
                        expired_date = datetime.strptime(
                            expired_date, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        raise ValueError(
                            "Invalid expired date format. required format is YYYY-MM-DD"
                        )
                    if expired_date < datetime.now().date():
                        raise ValueError(
                            "Expired date must be greater than current date."
                        )
                if stockin_type == "RECEIVE LENT MATERIAL":
                    lent_stock = BorrowedStock.objects.filter(
                        product=product, source=source, lent_quantity__gt=0
                    ).first()

                    if not lent_stock:
                        raise ValueError(
                            f"There is no lent material from source {source.name} for {product.name}"
                        )

                    if quantity > lent_stock.lent_quantity:
                        raise ValueError(
                            f"Receiving more quantity than lent for {product.name}. "
                            f"Actual lent quantity: {lent_stock.lent_quantity}"
                        )
                    lent_material_stock = lent_stock

            with transaction.atomic():
                quantity = quantity if product.is_mutli_type_unit else 1
                if stockin_type == "RETURN MATERIAL":
                    instance = barcode.stocks.first()
                    instance.quantity += quantity
                    instance.rack = rack
                    instance.save()

                elif is_used_stock_barcode:
                    instance = barcode.stocks.filter(rack=rack).first()
    
                    if instance:
                        instance.quantity += quantity
                        instance.created_at = datetime.now()
                        instance.save()
                    else:
                        instance = Stocks.objects.create(
                            quantity=quantity,
                            barcode=barcode,
                            rack=rack,
                            product=product,
                            expired_date=expired_date
                        )
                else:
                    instance = Stocks.objects.create(
                            quantity=quantity,
                            barcode=barcode,
                            rack=rack,
                            product=product,
                            expired_date=expired_date)

                if barcode.status != "Used":
                    barcode.status = "Used"
                    barcode.save()

                product.net_quantity += quantity
                product.save()

                stock_history = StocksHistory.objects.create(
                    stock=instance,
                    source=source,
                    history_type=stockin_type,
                    quantity=quantity,
                    product_quantity=product.net_quantity,
                    user=user,
                    is_stock_out=False,
                )

                if stockin_type == "BORROWING MATERIAL":
                    borrowed, _ = BorrowedStock.objects.get_or_create(
                        product=product, source=source
                    )
                    borrowed.borrowed_quantity += quantity
                    borrowed.save()

                if stockin_type == "RECEIVE LENT MATERIAL" and lent_material_stock:
                    lent_material_stock.lent_quantity -= quantity
                    lent_material_stock.save()

                if is_single_stock:
                    history_serializer = StocksHistorySerializer(stock_history)
                    return Response(
                        {
                            "success": True,
                            "message": "Stock IN Successfully",
                            "data": {
                                "product_net_quantity": product.net_quantity,
                                "rack": {
                                    "id": rack.id,
                                    "quantity": quantity,
                                    "rack_no": rack.rack_no,
                                },
                                "history": history_serializer.data,
                            },
                        },
                        status=200,
                    )

                return True
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "message": str(e)}, status=400)

    def send_low_stock_email_alert(self, products):
        try:
            email_to_list = Users.objects.filter(
                is_allow_email_low_stock_alert=True
            ).values_list("email", flat=True)
            if not email_to_list:
                return False
            template_path = "low_stock_alert.html"
            subject = "InteliStock Alert"
            for product in products:
                email_context = {
                    "product_name": product.name,
                    "current_stock": product.net_quantity,
                    "min_stock": product.min_threshold,
                    "product_url": f"{settings.FRONTEND_BASE_URL}/material/{product.id}",
                }
                sendEmail_template(
                    email_to=email_to_list,
                    subject=subject,
                    email_template_path=template_path,
                    email_context=email_context,
                )
        except Exception as e:
            Syserror(e)

    def validate_stock_out_barcode(self, request, barcode):
        try:
            if stocks := Stocks.objects.filter(barcode__barcode_no=barcode.upper(), quantity__gt=0).order_by('-quantity', 'created_at'):
                stock = stocks.first()
                rack_stock_list = []
                if stock.quantity > 0:
                    data = {
                        "id": stock.id,
                        "quantity": stock.quantity,
                        "product": stock.product.name,
                        "rack":stock.rack.rack_no,
                        "rack_stock_list":rack_stock_list
                    }
                    if stock.product.perishable_product:
                        older_stock_ids = request.GET.get("older_stock_id", "")
                        # Ensure older_stock_ids is a valid list
                        if older_stock_ids:
                            older_stock_ids = (
                                older_stock_ids.split(",")
                                if isinstance(older_stock_ids, str)
                                else []
                            )
                        else:
                            older_stock_ids = []
                        if (
                            lastStock := Stocks.objects.filter(
                                product=stock.product,
                                quantity__gt=0,
                                expired_date__isnull=False,
                            )
                            .exclude(id__in=older_stock_ids)
                            .order_by("expired_date")
                            .first()
                        ):
                            if lastStock != stock:
                                data["oldest_stock"] = {
                                    "id": lastStock.id,
                                    "quantity": lastStock.quantity,
                                    "barcode": lastStock.barcode.barcode_no,
                                    "rack": lastStock.rack.rack_no,
                                }
                            else:
                                print("old stock", lastStock)
                        else:
                            print("no stock found")
                    
                    if stock.product.is_mutli_type_unit and stocks.count() > 1:
                        data["rack_stock_list"] = [{"stock_id":s.id, "rack":s.rack.rack_no, "quantity":s.quantity} for s in stocks]
                    response = {
                        "success": True,
                        "message": "Stock Get Succesfully",
                        "data": data,
                    }
                    return Response(response, status=200)
                resp = {
                    "success": False,
                    "message": f"{barcode} Barcode have zero quantity",
                }
                return Response(resp, status=400)
            else:
                response = {
                    "success": False,
                    "message": "Invalid Barcode",
                }
                return Response(response, status=400)
            
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

class ProductsStockView(APIView, CustomPagination):
    serializer_class = ProductStocksSerializer

    def get(self, request, id=None):
        try:
            if id is None:
                response = {
                    "success": False,
                    "message": "required product Id to get Products Stock",
                }
                return Response(response, status=400)
            product = Products.objects.get(id=id)
            filter_criteria = Q(product=product, quantity__gt=0)
            if query := request.GET.get("query"):
                filter_criteria &= Q(
                    Q(rack__rack_no__icontains=query)
                    | Q(barcode__barcode_no__icontains=query)
                )
            instance = (
                product.stocks.select_related("rack", "barcode")
                .filter(filter_criteria)
                .order_by("expired_date", "created_at")
            )
            if export := request.GET.get("export") == "true":
                file_name = f"{product.name}-stocks-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_material_stocks_to_excel(instance, file_name)
            
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Products Stock List Get Succesfully",
                "data": data,
                "product": {"name": product.name, "id": id},
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        
    def export_material_stocks_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"

            # Headers
            headers = [
                "Barcode",
                "Rack",
                "Quantity",
                "Last Stock In Date"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.barcode.barcode_no,
                    row.rack.rack_no,
                    row.quantity,
                    row.updated_at.strftime("%d %b %Y, %I:%M %p"),
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)

    def put(self, request, id=None):
        try:
            if id is None:
                response = {
                    "success": False,
                    "message": "required product Id to update Stock Rack",
                }
                return Response(response, status=400)
            with transaction.atomic():
                data = request.data
                stock_id = data.get("stock_id", None)
                rack_id = data.get("rack_id", None)
                if not all([rack_id, stock_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                try:
                    stock = Stocks.objects.get(id=stock_id, product__id=id)
                except:
                    raise ValueError("Stock not found")

                try:
                    rack = Racks.objects.get(id=rack_id)
                except:
                    raise ValueError("Rack not found")
                
                if new_rack_stock := Stocks.objects.filter(rack=rack, barcode=stock.barcode, product__id=id, quantity__gt=0).first():
                    new_rack_stock.quantity = new_rack_stock.quantity + stock.quantity
                    new_rack_stock.save()
                    stock.quantity = 0
                    stock.save()
                else:
                    stock.rack = rack
                    stock.save()
                
                response = {
                    "success": True,
                    "message": "Stock Rack updated successfully",
                    "data": self.serializer_class(stock).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class PerishableStockView(APIView, CustomPagination):
    serializer_class = PerishableStocksSerializer

    def get(self, request):
        try:
            filter_criteria = Q(expired_date__isnull=False, quantity__gt=0)
            if query := request.GET.get("query"):
                filter_criteria &= Q(
                    Q(product__name__icontains=query)
                    | Q(source__name__icontains=query)
                    | Q(rack__rack_no__icontains=query)
                )
            
            now_date = datetime.today().date()
            expiry_type = request.GET.get("expiry_type", "Expiry Soon")
            if expiry_type == "Expiry Soon":
                soon_end_date = now_date + timedelta(days=60)
                filter_criteria &= Q(expired_date__range=[now_date , soon_end_date])
            else:
                filter_criteria &=Q(expired_date__lt=now_date)
                
            instance = (
                Stocks.objects.select_related("rack", "product", "barcode")
                .filter(filter_criteria)
                .order_by("expired_date")
            )

            if export := request.GET.get("export") == "true":
                file_name = f"{'Expiry-Soon' if expiry_type == 'Expiry Soon' else 'Expired'}-perishable-stock-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_perishable_stocks_to_excel(instance, file_name)
            
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Products List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        
    def export_perishable_stocks_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"

            # Headers
            headers = [
                "Material",
                "Barcode",
                "Rack",
                "Quantity",
                "Source",
                "Expiry Date"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.product.name,
                    row.barcode.barcode_no,
                    row.rack.rack_no,
                    row.quantity,
                    row.source.name,
                    row.expired_date.strftime("%d %b %Y"),
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)


class BorrowedStockView(APIView):

    def get(self, request, source_id=None):
        try:
            if source_id:
                if source := Source.objects.filter(id=source_id).first():
                    borrowed_products = (
                        BorrowedStock.objects.filter(source=source)
                        .values("product__name")
                        .annotate(
                            total_borrowed_quantity=Sum("borrowed_quantity"),
                            total_lent_quantity=Sum("lent_quantity"),
                        )
                        .filter(
                            Q(total_borrowed_quantity__gt=0)
                            | Q(total_lent_quantity__gt=0)
                        )  # Filter non-zero values
                        .order_by("product__name")
                    )

                    response = {
                        "success": True,
                        "message": "Borrowed Stock Get Succesfully",
                        "data": borrowed_products,
                    }
                    return Response(response, status=200)
                else:
                    response = {
                        "success": False,
                        "message": "Source Not Found",
                    }
                    return Response(response, status=400)
            # list details of all borrowed & lent stock
            instance = (
                BorrowedStock.objects.filter(
                    Q(borrowed_quantity__gt=0) | Q(lent_quantity__gt=0)
                )
                .values("source__id", "source__name")  # Group by source
                .annotate(
                    product_count=Count(
                        "product",
                        filter=Q(borrowed_quantity__gt=0) | Q(lent_quantity__gt=0),
                        distinct=True,
                    ),  # Count distinct products
                    total_borrowed_quantity=Sum("borrowed_quantity"),  # Sum borrowed
                    total_lent_quantity=Sum("lent_quantity"),  # Sum lent
                )
                .distinct()  # Ensure unique sources
                .order_by("source__id")  # Order by source ID
            )
            response = {
                "success": True,
                "message": "Borrowed Stock List Get Succesfully",
                "data": instance,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class StockHistoryView(APIView, CustomPagination):
    serializer_class = StocksHistorySerializer

    def get(self, request, id=None):
        try:
            if id:
                if product := Products.objects.filter(id=id).first():
                    f = Q(stock__product__id=id)
                    if query := request.GET.get("query"):
                        f &= Q(
                            Q(stock__rack__rack_no__icontains=query)
                            | Q(stock__barcode__barcode_no__icontains=query)
                            | Q(quantity__icontains=query)
                            | Q(source__name__icontains=query)
                            | Q(purpose__icontains=query)
                            | Q(user__name__icontains=query)
                        )
                    if history_type := request.GET.get("history_type"):
                        f &= Q(history_type=history_type)
                    instance = (
                        StocksHistory.objects.select_related(
                            "stock",
                            "stock__rack",
                            "stock__barcode",
                            "stock__product",
                            "source",
                            "employee",
                            "user",
                        )
                        .filter(f)
                        .order_by("-created_at")
                    )
                    if page := self.paginate_queryset(instance, request, view=self):
                        serializer = self.serializer_class(page, many=True)
                        result = self.get_paginated_response(serializer.data)
                        data = result.data["results"]  # pagination data
                        total = result.data["count"]
                        response = {
                            "success": True,
                            "message": "Stock History List Get Succesfully",
                            "data": data,
                            "product": {"name": product.name, "id": id},
                            "total": total,
                        }
                        return Response(response, status=200)
                else:
                    response = {
                        "success": False,
                        "message": "Product Not Found",
                    }
                    return Response(response, status=400)
            
            filter = Q()
            if query := request.GET.get("query"):
                filter &= Q(
                    Q(stock__rack__rack_no__icontains=query)
                    | Q(stock__product__name__icontains=query)
                    | Q(stock__product__ucs_code__icontains=query)
                    | Q(stock__barcode__barcode_no__icontains=query)
                    | Q(source__name__icontains=query)
                    | Q(user__name__icontains=query)
                    | Q(user__email__icontains=query)
                    | Q(employee__personnel_number__icontains=query)
                    | Q(employee__name__icontains=query)
                    | Q(purpose__icontains=query)
                )

            if from_date:= request.GET.get("from_date", ""):
                try:
                    from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                    filter &= Q(created_at__date__gte=from_date)
                except:
                    pass
            
            if to_date:= request.GET.get("to_date", ""):
                try:
                    to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
                    filter &= Q(created_at__date__lte=to_date)
                except:
                    pass
            
            if history_status := request.GET.get("status"):
                if history_status == "In":
                    filter &= Q(is_stock_out=False)
                elif history_status == "Out":
                    filter &= Q(is_stock_out=True)
            
            if history_type := request.GET.get("history_type"):
                filter &= Q(history_type=history_type)
            
            instance = (
                StocksHistory.objects.select_related(
                    "stock",
                    "stock__rack",
                    "stock__barcode",
                    "stock__product",
                    "source",
                    "employee",
                    "user",
                )
                .filter(filter)
                .order_by("-created_at")
            )
            if export := request.GET.get("export") == "true":
                file_name = f"stock-history-details-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_stock_history_to_excel(instance, file_name)

            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()

            response = {
                "success": True,
                "message": "Stock History List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        

    def export_stock_history_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"

            # Headers
            headers = [
                "Date",
                "Material",
                "UCS Code",
                "Rack",
                "Quantity",
                "Barcode",
                "Status",
                "Status Type",
                "Perform By",
                "Source",
                "Employee",
                "Purpose"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.created_at.strftime("%d %b %Y, %I:%M %p"),
                    row.stock.product.name,
                    row.stock.product.ucs_code,
                    row.stock.rack.rack_no,
                    row.quantity,
                    row.stock.barcode.barcode_no,
                    f"{'STOCK OUT' if row.is_stock_out else 'STOCK IN' }",
                    row.history_type,
                    row.user.name,
                    row.source.name if row.source else '',
                    f"{row.employee.name if row.employee else ''}",
                    f"{row.purpose or ''}"
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)


class BarcodeStockHistoryView(APIView, CustomPagination):
    serializer_class = StocksHistorySerializer

    def get(self, request, barcode=None):
        try:
            if not barcode:
                response = {
                    "success": False,
                    "message": "Required Material Barcode",
                }
                return Response(response, status=400)
            
            barcode = Barcodes.objects.filter(barcode_no=barcode.upper(), is_product_type=True).first()
            if not barcode:
                response = {
                    "success": False,
                    "message": "Not a Valid Material barcode",
                }
                return Response(response, status=400)
            
            f = Q(stock__barcode=barcode)
            if query := request.GET.get("query"):
                f &= Q(
                    Q(stock__rack__rack_no__icontains=query)
                    | Q(quantity__icontains=query)
                    | Q(source__name__icontains=query)
                    | Q(purpose__icontains=query)
                    | Q(user__name__icontains=query)
                )
            if history_type := request.GET.get("history_type"):
                f &= Q(history_type=history_type)
            instance = (
                StocksHistory.objects.select_related(
                    "stock",
                    "stock__rack",
                    "stock__barcode",
                    "stock__product",
                    "source",
                    "employee",
                    "user",
                )
                .filter(f)
                .order_by("-created_at")
            )

            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]
                response = {
                    "success": True,
                    "message": "Barcode Stock History List Get Succesfully",
                    "data": data,
                    "total": total,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

class BarcodeView(APIView, CustomPagination):
    serializer_class = BarcodeSerializer

    def get(self, request):
        try:
            product_barcode_no = request.GET.get("product_barcode_no", None)
            product_id = request.GET.get("product_id", None)
            if product_barcode_no and product_id:
                return self.validate_barcode(product_barcode_no, product_id)

            elif return_stock_barcode_no := request.GET.get(
                "return_stock_barcode_no", None
            ):
                return self.validate_return_stock_barcode(return_stock_barcode_no)

            filter = Q()
            if query := request.GET.get("query"):
                filter &= Q(
                    Q(barcode_no=query.upper())
                    | Q(racks__rack_no__icontains=query)
                    | Q(stocks__product__name__icontains=query)
                )
            if barcode_status := request.GET.get("status"):
                filter &= (
                    Q(status__exact=barcode_status.capitalize())
                    if barcode_status in ["Used", "Unused"]
                    else Q()
                )

            instance = (
                Barcodes.objects.select_related("racks")
                .filter(filter)
                .order_by("-created_at")
            )
            if export := request.GET.get("export") == "true":
                file_name = f"barcode-data-{datetime.now().strftime('%d%m%y-%H%M')}.xlsx"
                return self.export_barcode_to_excel(instance, file_name)

            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Barcode List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)
        
    def export_barcode_to_excel(self, data, file_name):
        # print(buses_map)
        """
        Export the bus occupancy data (buses_map) to Excel with grouped bus data.
        """
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "sheet 1"
            # Headers
            headers = [
                "Barcode",
                "Status",
                "Created At"
            ]
            sheet.append(headers)

            for row in data:
                sheet.append([
                    row.barcode_no,
                    row.status,
                    row.created_at.strftime("%d %b %Y, %I:%M %p")
                ])

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # Write to output stream
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            # Prepare response
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        
        except Exception as e:
            Syserror(e)
            return Response({"success": False, "detail": str(e)}, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                barcode_type = data.get("barcode_type", "rack")
                if no_of_barcode := data.get("no_of_barcode", 1):
                    no_of_barcode = (
                        int(no_of_barcode)
                        if isinstance(no_of_barcode, str)
                        else no_of_barcode
                    )

                if not all([barcode_type, no_of_barcode]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)

                if no_of_barcode > 21:
                    response = {
                        "success": False,
                        "message": "Maximum No of barcode generation upto 20 20",
                    }
                    return Response(response, status=400)

                is_product_type = barcode_type == "product"
                barcode_ids = []
                for _ in range(1, no_of_barcode + 1):
                    barcode = Barcodes.objects.create(
                        is_product_type=is_product_type, status="Unused"
                    )
                    barcode_ids.append(barcode.id)

                instance = Barcodes.objects.filter(id__in=barcode_ids)
                response = {
                    "success": True,
                    "message": "Barcode created  successfully",
                    "data": self.serializer_class(instance, many=True).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def validate_barcode(self, barcode_no, product_id):
        barcode = Barcodes.objects.filter(barcode_no=barcode_no.upper()).first()
        if not barcode:
            response = {
                "success": False,
                "message": "Invalid barcode",
            }
            return Response(response, status=400)

        if not barcode.is_product_type:
            response = {
                "success": False,
                "message": "Not a Material barcode",
            }
            return Response(response, status=400)

        product = Products.objects.filter(id=product_id).first()
        if not product:
            response = {
                "success": False,
                "message": "Validations failed, required product id",
            }
            return Response(response, status=400)
        
        if product.is_mutli_type_unit:
            if barcode.status == "Used":

                if product.perishable_product:
                    response = {
                        "success": False,
                        "message": "Barcode already used, Can not accept duplicate Barcode for perishable material",
                    }
                    return Response(response, status=400)
                
                barcode_all_stock = barcode.stocks.all()
                if barcode_all_stock.exclude(product__id=product_id).count() != 0:
                    response = {
                        "success": False,
                        "message": "Barcode already used in another material.",
                    }
                    return Response(response, status=400)
                
        elif barcode.status == "Used":
                response = {
                    "success": False,
                    "message": "Barcode already used, Can not accept duplicate Barcode for single unit type material",
                }
                return Response(response, status=400)

        response = {
            "success": True,
            "message": "Valid Material Barcode",
        }
        return Response(response, status=200)

    def validate_return_stock_barcode(self, barcode_no):
        barcode = Barcodes.objects.filter(barcode_no=barcode_no.upper()).first()
        if not barcode:
            response = {
                "success": False,
                "message": "Barcode not found",
            }
            return Response(response, status=400)

        if not barcode.is_product_type:
            response = {
                "success": False,
                "message": "Not a Material barcode",
            }
            return Response(response, status=400)

        if barcode.status != "Used":
            response = {
                "success": False,
                "message": "Barcode is not used yet, impossible receive to return material",
            }
            return Response(response, status=400)

        if not barcode.stocks:
            response = {
                "success": False,
                "message": "Stock not found",
            }
            return Response(response, status=400)
        
        stock = barcode.stocks.first()
        histoty = stock.stockshistory_set.filter(
            history_type="MATERIAL CONSUMPTION", is_stock_out=True
        )
        if not histoty.exists():
            response = {
                "success": False,
                "message": "Stock history not found for MATERIAL CONSUMPTION type",
            }
            return Response(response, status=400)

        if stock.product.is_mutli_type_unit:
            quantity = histoty.aggregate(Sum("quantity"))["quantity__sum"]
        else:
            quantity = 1

        data = {
            "product": {
                "value": stock.product.id,
                "label": stock.product.name,
                "is_multi_unit": stock.product.is_mutli_type_unit,
            },
            "quantity": quantity,
            "rack": {"value": stock.rack.id, "label": stock.rack.rack_no},
        }

        response = {
            "success": True,
            "message": "Valid Material Barcode",
            "data": data,
        }
        return Response(response, status=200)


# front end search part


class BarcodeActionView(APIView):
    def get(self, request, barcode):
        try:
            if not barcode:
                response = {
                    "success": False,
                    "message": "Barcode required",
                }
                return Response(response, status=400)

            barcode = Barcodes.objects.filter(barcode_no=barcode.upper()).first()
            if not barcode:
                response = {
                    "success": False,
                    "message": "Barcode  not exists",
                }
                return Response(response, status=400)
            action = []
            if barcode.is_product_type:
                if barcode.status == "Unused":
                    action = ["STOCKIN"]
                else:
                    if stock := barcode.stocks.first():
                        action = (
                            ["STOCKOUT", "HISTORY"]
                            if stock.quantity > 0
                            else ["HISTORY"]
                        )
                    else:
                        response = {"success": False, "message": "Unkown error"}
                        return Response(response, status=400)
            else:
                action = ["RACK"]
            response = {
                "success": True,
                "message": "Barcode Action Get Succesfully",
                "action": action,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchRacksView(APIView):
    serializer_class = SearchRacksSerializer

    def get(self, request):
        try:
            instance = Racks.objects.select_related().order_by("rack_no")
            serializer = self.serializer_class(instance, many=True)
            data = serializer.data
            response = {
                "success": True,
                "message": "Rack List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchSourceView(APIView):
    serializer_class = SourceSerializer

    def get(self, request):
        try:
            stock_type = request.GET.get("stock_type", "NEW MATERIAL")
            is_central_store = stock_type in [
                "NEW MATERIAL",
                "RETURN MATERIAL",
                "MATERIAL CONSUMPTION",
            ]
            filter_criteria = Q(is_central_store=is_central_store)
            if query := request.GET.get("query"):
                filter_criteria &= Q(name__icontains=query)

            instance = Source.objects.filter(filter_criteria).order_by("name")
            serializer = self.serializer_class(instance, many=True)
            response = {
                "success": True,
                "message": "Source List Get Succesfully",
                "data": serializer.data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchPurposeView(APIView):
    def get(self, request):
        try:
            instance = (
                StocksHistory.objects.filter(
                    is_stock_out=True, history_type="MATERIAL CONSUMPTION"
                )
                .values("purpose")
                .distinct()
                .order_by("purpose")
            )
            response = {
                "success": True,
                "message": "Purpose List  Succesfully",
                "data": [
                    {"value": d["purpose"], "label": d["purpose"]} for d in instance
                ],
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchMaterialView(APIView):
    serializer_class = SearchMaterialSerializer

    def get(self, request):
        try:
            instance = Products.objects.filter(is_active=True).only("name", "id").order_by("name")
            response = {
                "success": True,
                "message": "Material List  Succesfully",
                "data": SearchMaterialSerializer(instance, many=True).data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchProductView(APIView):
    def get(self, request):
        try:
            if query := request.GET.get("query"):
                results = (
                    Products.objects.select_related()
                    .filter(
                        Q(name__icontains=query)
                        | Q(stocks__barcode__barcode_no__icontains=query)
                        | Q(ucs_code__icontains=query)
                    )
                    .values("name", "id")
                    .distinct()
                    .order_by("name")
                )
            else:
                results = []
            response = {
                "success": True,
                "message": "Search Result",
                "data": results,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SearchEmployeeView(APIView):
    serializer_class = SearchEmployeeSerializer

    def get(self, request):
        try:
            instance = Employees.objects.order_by("name")
            serializer = self.serializer_class(instance, many=True)
            data = serializer.data
            response = {
                "success": True,
                "message": "Employee List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class DesignationView(APIView):
    serializer_class = DesignationSerializer

    def get(self, request):
        try:
            if query := request.GET.get("query"):
                instance = (
                    Designation.objects.select_related()
                    .filter(name__icontains=query)
                    .order_by("name")
                )
            else:
                instance = Designation.objects.all().order_by("name")
            serializer = self.serializer_class(instance, many=True)
            response = {
                "success": True,
                "message": "Designation List Get Succesfully",
                "data": serializer.data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                name = data.get("name", None)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required designation Name",
                    }
                    return Response(response, status=400)

                if Designation.objects.filter(name=name).exists():
                    response = {
                        "success": False,
                        "message": "Designation already exists",
                    }
                    return Response(response, status=400)

                instance = Designation.objects.create(name=name)
                response = {
                    "success": True,
                    "message": "Designation Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                designation_id = data.get("designation_id", None)
                name = data.get("name", None)
                if not all([name, designation_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                instance = Designation.objects.get(id=designation_id)

                if (
                    Designation.objects.filter(name=name)
                    .exclude(name=instance.name)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Designation already exists",
                    }
                    return Response(response, status=400)
                instance.name = name
                instance.save()
                response = {
                    "success": True,
                    "message": "Designation Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class UnitView(APIView):
    serializer_class = UnitSerializer

    def get(self, request):
        try:
            if query := request.GET.get("query"):
                instance = (
                    Unit.objects.select_related()
                    .filter(name__icontains=query)
                    .order_by("name")
                )
            else:
                instance = Unit.objects.all().order_by("name")
            serializer = self.serializer_class(instance, many=True)
            response = {
                "success": True,
                "message": "Unit List Get Succesfully",
                "data": serializer.data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                name = data.get("name", None)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required unit Name",
                    }
                    return Response(response, status=400)

                if Unit.objects.filter(name=name).exists():
                    response = {
                        "success": False,
                        "message": "Unit already exists",
                    }
                    return Response(response, status=400)

                instance = Unit.objects.create(name=name)
                response = {
                    "success": True,
                    "message": "Unit Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                unit_id = data.get("unit_id", None)
                name = data.get("name", None)
                if not all([name, unit_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                instance = Unit.objects.get(id=unit_id)

                if Unit.objects.filter(name=name).exclude(name=instance.name).exists():
                    response = {
                        "success": False,
                        "message": "Unit already exists",
                    }
                    return Response(response, status=400)
                instance.name = name
                instance.save()
                response = {
                    "success": True,
                    "message": "Unit Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class SourceView(APIView):
    serializer_class = SourceSerializer

    def get(self, request):
        try:
            if query := request.GET.get("query"):
                instance = Source.objects.filter(name__icontains=query).order_by("name")
            else:
                instance = Source.objects.all().order_by("name")
            serializer = self.serializer_class(instance, many=True)
            response = {
                "success": True,
                "message": "Source List Get Succesfully",
                "data": serializer.data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                name = data.get("name", None)
                is_central_store = data.get("central_store", False)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required Source Name",
                    }
                    return Response(response, status=400)

                if Source.objects.filter(name=name).exists():
                    response = {
                        "success": False,
                        "message": "Source already exists",
                    }
                    return Response(response, status=400)

                instance = Source.objects.create(
                    name=name, is_central_store=is_central_store
                )
                response = {
                    "success": True,
                    "message": "Source Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                source_id = data.get("source_id", None)
                name = data.get("name", None)
                is_central_store = data.get("central_store", False)
                if not all([name, source_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                instance = Source.objects.get(id=source_id)

                if (
                    Source.objects.filter(name=name)
                    .exclude(name=instance.name)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Source already exists",
                    }
                    return Response(response, status=400)
                instance.name = name
                instance.is_central_store = is_central_store
                instance.save()
                response = {
                    "success": True,
                    "message": "Source Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class EmployeeView(APIView, CustomPagination):
    serializer_class = EmployeeSerializer

    def get(self, request):
        try:
            if query := request.GET.get("query"):
                instance = (
                    Employees.objects.select_related()
                    .filter(
                        Q(name__icontains=query)
                        | Q(personnel_number__icontains=query)
                        | Q(phone__icontains=query)
                    )
                    .annotate(
                        product=Count("stockshistory__stock__product", distinct=True)
                    )
                    .order_by("name")
                )
            else:
                instance = (
                    Employees.objects.all()
                    .annotate(
                        product=Count("stockshistory__stock__product", distinct=True)
                    )
                    .order_by("name")
                )
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "Employee List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                personnel_number = data.get("personnel_number", None)
                name = data.get("name", None)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required name fields",
                    }
                    return Response(response, status=400)

                if (
                    personnel_number
                    and Employees.objects.filter(
                        personnel_number=personnel_number
                    ).exists()
                ):
                    response = {
                        "success": False,
                        "message": f"Employee With personnel number {personnel_number} already exists",
                    }
                    return Response(response, status=400)

                instance = Employees.objects.create(
                    name=name,
                    personnel_number=personnel_number,
                    phone=data.get("phone", None),
                )
                response = {
                    "success": True,
                    "message": "Employee Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                employee_id = data.get("employee_id", None)
                name = data.get("name", None)
                personnel_number = data.get("personnel_number", None)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required name Field",
                    }
                    return Response(response, status=400)
                instance = Employees.objects.get(id=employee_id)

                if personnel_number and (
                    Employees.objects.filter(personnel_number=personnel_number)
                    .exclude(personnel_number=instance.personnel_number)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": f"Employee With personnel number {personnel_number} already exists",
                    }
                    return Response(response, status=400)
                instance.name = name
                instance.personnel_number = personnel_number
                instance.phone = data.get("phone", None)
                instance.save()
                response = {
                    "success": True,
                    "message": "Employee Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def delete(self, request, id):
        try:
            with transaction.atomic():
                instance = Employees.objects.get(id=id)
                response = {
                    "success": True,
                    "message": "Employee Deleted Successfully",
                }
                instance.delete()
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ProductCategoryView(APIView):
    serializer_class = ProductCategorySerializer

    def get(self, request):
        try:
            if query := request.GET.get("query"):
                instance = (
                    ProductCategory.objects.select_related()
                    .filter(name__icontains=query)
                    .order_by("name")
                )
            else:
                instance = ProductCategory.objects.all().order_by("name")
            serializer = self.serializer_class(instance, many=True)
            response = {
                "success": True,
                "message": "Product Category List Get Succesfully",
                "data": serializer.data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            with transaction.atomic():
                data = request.data
                name = data.get("name", None)
                if not name:
                    response = {
                        "success": False,
                        "message": "Required designation Name",
                    }
                    return Response(response, status=400)

                if ProductCategory.objects.filter(name=name).exists():
                    response = {
                        "success": False,
                        "message": "Product Category already exists",
                    }
                    return Response(response, status=400)

                instance = ProductCategory.objects.create(name=name)
                response = {
                    "success": True,
                    "message": "Product Category Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        try:
            with transaction.atomic():
                data = request.data
                product_category_id = data.get("product_category_id", None)
                name = data.get("name", None)
                if not all([name, product_category_id]):
                    response = {
                        "success": False,
                        "message": "Required All Field",
                    }
                    return Response(response, status=400)
                instance = ProductCategory.objects.get(id=product_category_id)

                if (
                    ProductCategory.objects.filter(name=name)
                    .exclude(name=instance.name)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Designation already exists",
                    }
                    return Response(response, status=400)
                instance.name = name
                instance.save()
                response = {
                    "success": True,
                    "message": "Product Category Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)


class UserView(APIView, CustomPagination):
    serializer_class = UserSerializer

    def get(self, request, id=None):
        if request.user.is_superuser is False:
            response = {"success": False, "message": "User not allowed"}
            return Response(response, status=400)
        try:
            if id:
                instance = Users.objects.filter(id=id, is_superuser=False)
                serializer = self.serializer_class(instance)
                response = {
                    "success": True,
                    "message": "Employee Get Succesfully",
                    "data": serializer.data,
                }
                return Response(response, status=200)
            if query := request.GET.get("query"):
                instance = (
                    Users.objects.select_related()
                    .filter(
                        Q(is_superuser=False),
                        Q(name__icontains=query)
                        | Q(email__icontains=query)
                        | Q(personnel_number__icontains=query)
                        | Q(designation__name__icontains=query)
                        | Q(mobile_number=query),
                    )
                    .order_by("-created_at")
                )
            else:
                instance = (
                    Users.objects.select_related()
                    .filter(is_superuser=False)
                    .order_by("-created_at")
                )
            if page := self.paginate_queryset(instance, request, view=self):
                serializer = self.serializer_class(page, many=True)
                result = self.get_paginated_response(serializer.data)
                data = result.data["results"]  # pagination data
                total = result.data["count"]  # pagination data
            else:
                serializer = self.serializer_class(instance, many=True)
                data = serializer.data
                total = instance.count()
            response = {
                "success": True,
                "message": "User List Get Succesfully",
                "data": data,
                "total": total,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def post(self, request):
        try:
            if request.user.is_superuser is False:
                response = {"success": False, "message": "User not allowed"}
                return Response(response, status=400)
            with transaction.atomic():
                data = request.data
                name = data.get("name", None)
                password = data.get("password", None)
                email = data.get("email", None)
                mobile_number = data.get("mobile_number", None)
                designation_id = data.get("designation", None)
                personnel_number = data.get("personnel_number", None)
                is_allow_email_daily_report = data.get(
                    "is_allow_email_daily_report", False
                )
                is_allow_email_low_stock_alert = data.get(
                    "is_allow_email_low_stock_alert", False
                )
                user_permission = data.get("permission")
                if not all(
                    [
                        name,
                        password,
                        email,
                        mobile_number,
                        designation_id,
                        personnel_number,
                    ]
                ):
                    response = {
                        "success": False,
                        "message": "All fields are required.",
                    }
                    return Response(response, status=400)

                if not isinstance(user_permission, dict):
                    response = {
                        "success": False,
                        "message": "Invalid permision data, required object",
                    }
                    return Response(response, status=400)
                invalid_keys = [
                    key for key in user_permission.keys() if key not in PERMISSION_KEYS
                ]
                if invalid_keys:
                    response = {
                        "success": False,
                        "message": f"Unknown permissions detected: {invalid_keys}",
                    }
                    return Response(response, status=400)
                phone_pattern = re.compile(r"\b\d{10}\b")
                email_pattern = re.compile(
                    r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b"
                )
                if not email_pattern.match(email):
                    response = {
                        "success": False,
                        "message": "Invalid Email Format",
                    }
                    return Response(response, status=400)
                if not phone_pattern.match(mobile_number):
                    response = {
                        "success": False,
                        "message": "Invalid Mobile Number Format",
                    }
                    return Response(response, status=400)
                designation = Designation.objects.filter(id=designation_id).first()
                if not designation:
                    response = {
                        "success": False,
                        "message": "Invalid designation",
                    }
                    return Response(response, status=400)
                if Users.objects.filter(email=email).exists():
                    response = {
                        "success": False,
                        "message": "Email Already Exists",
                    }
                    return Response(response, status=400)
                if user_permission["manage_stockout"]:
                    user_permission["consumption_stockout"] = False
                instance = Users.objects.create_user(
                    name=name,
                    email=email,
                    password=password,
                    mobile_number=mobile_number,
                    designation=designation,
                    personnel_number=personnel_number,
                    is_allow_email_daily_report=is_allow_email_daily_report,
                    is_allow_email_low_stock_alert=is_allow_email_low_stock_alert,
                    is_superuser=False,
                    user_permission=user_permission,
                    is_staff=True,
                    is_active=True,
                    role="User",
                )
                response = {
                    "success": True,
                    "message": "User Created  Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def put(self, request):
        if request.user.is_superuser is False:
            response = {"success": False, "message": "User not allowed"}
            return Response(response, status=400)
        try:
            with transaction.atomic():
                data = request.data
                emp_id = data.get("emp_id", None)
                name = data.get("name", None)
                password = data.get("password", None)
                email = data.get("email", None)
                status = data.get("status", "Active")
                mobile_number = data.get("mobile_number", None)
                designation_id = data.get("designation", None)
                personnel_number = data.get("personnel_number", None)
                is_allow_email_daily_report = data.get(
                    "is_allow_email_daily_report", False
                )
                is_allow_email_low_stock_alert = data.get(
                    "is_allow_email_low_stock_alert", False
                )
                user_permission = data.get("permission")
                if not all(
                    [
                        emp_id,
                        name,
                        email,
                        mobile_number,
                        designation_id,
                        personnel_number,
                    ]
                ):
                    response = {
                        "success": False,
                        "message": "Required All fields",
                    }
                    return Response(response, status=400)
                if not isinstance(user_permission, dict):
                    response = {
                        "success": False,
                        "message": "Invalid permision data, required object",
                    }
                    return Response(response, status=400)
                invalid_keys = [
                    key for key in user_permission.keys() if key not in PERMISSION_KEYS
                ]
                if invalid_keys:
                    response = {
                        "success": False,
                        "message": f"Unknown permissions detected: {invalid_keys}",
                    }
                    return Response(response, status=400)

                phone_pattern = re.compile(r"\b\d{10}\b")
                email_pattern = re.compile(
                    r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b"
                )
                if not email_pattern.match(email):
                    response = {
                        "success": False,
                        "message": "Invalid Email Format",
                    }
                    return Response(response, status=400)
                if not phone_pattern.match(mobile_number):
                    response = {
                        "success": False,
                        "message": "Invalid Mobile Number Format",
                    }
                    return Response(response, status=400)

                designation = Designation.objects.filter(id=designation_id).first()
                if not designation:
                    response = {
                        "success": False,
                        "message": "Invalid designation",
                    }
                    return Response(response, status=400)
                instance = Users.objects.get(id=emp_id)
                if (
                    Users.objects.filter(email=email)
                    .exclude(email=instance.email)
                    .exists()
                ):
                    response = {
                        "success": False,
                        "message": "Email Already Exists",
                    }
                    return Response(response, status=400)
                if (
                    instance.personnel_number != personnel_number
                    or instance.name != name
                    or instance.mobile_number != mobile_number
                ):
                    try:
                        employee = Employees.objects.get(
                            personnel_number=instance.personnel_number
                        )
                        employee.name = name
                        employee.personnel_number = personnel_number
                        employee.phone = mobile_number
                        employee.save()
                    except:
                        pass
                if user_permission["manage_stockout"]:
                    user_permission["consumption_stockout"] = False
                instance.name = name
                instance.designation = designation
                instance.email = email
                instance.mobile_number = mobile_number
                instance.personnel_number = personnel_number
                instance.user_permission = user_permission
                instance.is_active = status == "Active"
                instance.is_allow_email_daily_report = is_allow_email_daily_report
                instance.is_allow_email_low_stock_alert = is_allow_email_low_stock_alert
                if password:
                    instance.set_password(password)
                instance.save()
                response = {
                    "success": True,
                    "message": "User Updated Successfully",
                    "data": self.serializer_class(instance).data,
                }
                return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {
                "success": False,
                "message": str(e),
            }
            return Response(response, status=400)

    def delete(self, request, id):
        if request.user.is_superuser is False:
            response = {"success": False, "message": "User not allowed"}
            return Response(response, status=400)
        try:
            with transaction.atomic():
                instance = Users.objects.filter(id=id, is_superuser=False)
                if instance.exists():
                    response = {
                        "success": True,
                        "message": "User Deleted Successfully",
                    }
                    instance.delete()
                    return Response(response, status=200)
                response = {
                    "success": False,
                    "message": f"User Not Found with This ID {id}",
                }
                return Response(response, status=400)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ABCAnalysisView(APIView):
    def get(self, request):
        try:
            f = Q(is_stock_out=True)
            view_type = request.GET.get("view_type")
            if view_type == "30days":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=31)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)
            elif view_type == "1year":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=366)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)
            stock_history = StocksHistory.objects.select_related(
                "rack", "stock"
            ).filter(f)
            total_quantity = stock_history.aggregate(total=Sum("quantity"))["total"]
            all_products = (
                stock_history.values("stock__product__name")
                .annotate(sum_quantity=Sum("quantity"))
                .order_by("-sum_quantity")
            )

            a_limit = 0.4 * total_quantity
            b_limit = 0.8 * total_quantity
            cumulative_quantity = 0
            A = 0
            B = 0
            C = 0
            product_data = {"A": [], "B": [], "C": []}
            for product in all_products:
                percent = round((product["sum_quantity"] / total_quantity) * 100, 2)
                if cumulative_quantity <= a_limit:
                    A += 1
                    product_data["A"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                elif cumulative_quantity <= b_limit:
                    B += 1
                    product_data["B"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                else:
                    product_data["C"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                    C += 1
                cumulative_quantity += product["sum_quantity"]

            labels = ["A", "B", "C"]
            output_data = [A, B, C]
            data = {"labels": labels, "data": output_data, "product_data": product_data}

            response = {
                "success": True,
                "message": "Product Forecast",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class VEDAnalysisView(APIView):
    def get(self, request):
        try:
            all_products = Products.objects.all().order_by("-net_quantity")
            prducts = all_products.aggregate(
                vital=Count("id", filter=Q(ved_category="Vital")),
                essential=Count("id", filter=Q(ved_category="Essential")),
                desirable=Count("id", filter=Q(ved_category="Desirable")),
            )
            labels = ["Vital", "Essential", "Desirable"]
            output_data = [prducts["vital"], prducts["essential"], prducts["desirable"]]
            data = {
                "labels": labels,
                "data": output_data,
                "product_data": all_products.values(
                    "name", "ved_category", "net_quantity"
                ),
            }
            response = {
                "success": True,
                "message": "Product Forecast",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class HMLAnalysisView(APIView):
    def get(self, request):
        try:
            f = Q(is_stock_out=True)
            view_type = request.GET.get("view_type")
            if view_type == "30days":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=31)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)
            elif view_type == "1year":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=366)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)

            stock_history = StocksHistory.objects.select_related(
                "rack", "stock"
            ).filter(f)
            total_quantity = stock_history.aggregate(
                total=Sum(F("quantity") * F("stock__product__price"))
            )["total"]
            all_products = (
                stock_history.values("stock__product__name")
                .annotate(sum_price=Sum(F("quantity") * F("stock__product__price")))
                .order_by("-sum_price")
            )
            h_limit = 0.2 * total_quantity
            m_limit = 0.3 * total_quantity
            cumulative_quantity = 0
            H = 0
            M = 0
            L = 0
            product_data = {"H": [], "M": [], "L": []}
            for product in all_products:
                percent = round((product["sum_price"] / total_quantity) * 100, 2)
                if cumulative_quantity <= h_limit:
                    H += 1
                    product_data["H"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "price": product["sum_price"],
                        }
                    )
                elif cumulative_quantity <= m_limit:
                    M += 1
                    product_data["M"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "price": product["sum_price"],
                        }
                    )
                else:
                    L += 1
                    product_data["L"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "price": product["sum_price"],
                        }
                    )
                cumulative_quantity += product["sum_price"]

            labels = ["H", "M", "L"]
            output_data = [H, M, L]
            data = {"labels": labels, "data": output_data, "product_data": product_data}
            response = {
                "success": True,
                "message": "Product Forecast",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class FSNAnalysisView(APIView):
    def get(self, request):
        try:
            f = Q(is_stock_out=True)
            view_type = request.GET.get("view_type")
            if view_type == "30days":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=31)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)
            elif view_type == "1year":
                from_date = datetime.now().date()
                to_date = from_date - timedelta(days=366)
                f &= Q(created_at__gte=to_date, created_at__lte=from_date)
            stock_history = StocksHistory.objects.select_related(
                "rack", "stock"
            ).filter(f)
            total_quantity = stock_history.aggregate(total=Sum("quantity"))["total"]
            all_products = (
                stock_history.values("stock__product__name")
                .annotate(sum_quantity=Sum("quantity"))
                .order_by("-sum_quantity")
            )

            f_limit = 0.15 * total_quantity
            s_limit = 0.35 * total_quantity
            cumulative_quantity = 0
            F = 0
            S = 0
            N = 0
            product_data = {"F": [], "S": [], "N": []}
            for product in all_products:
                percent = round((product["sum_quantity"] / total_quantity) * 100, 2)
                if cumulative_quantity <= f_limit:
                    F += 1
                    product_data["F"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                elif cumulative_quantity <= s_limit:
                    S += 1
                    product_data["S"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                else:
                    N += 1
                    product_data["N"].append(
                        {
                            "name": product["stock__product__name"],
                            "percentage": percent,
                            "quantity": product["sum_quantity"],
                        }
                    )
                cumulative_quantity += product["sum_quantity"]

            labels = ["F", "S", "N"]
            output_data = [F, S, N]
            data = {"labels": labels, "data": output_data, "product_data": product_data}

            response = {
                "success": True,
                "message": "Product Forecast",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class SDEAnalysisView(APIView):
    def get(self, request):
        try:
            products = Products.objects.all()
            total_lead_time = products.aggregate(total=Sum("lead_time"))["total"]
            all_products = products.values("name", "lead_time").order_by("-lead_time")

            s_limit = 0.2 * total_lead_time
            d_limit = 0.3 * total_lead_time
            cumulative_quantity = 0
            S = 0
            D = 0
            E = 0
            product_data = {"S": [], "D": [], "E": []}
            for product in all_products:
                percent = round((product["lead_time"] / total_lead_time) * 100, 2)
                if cumulative_quantity <= s_limit:
                    S += 1
                    product_data["S"].append(
                        {
                            "name": product["name"],
                            "percentage": percent,
                            "lead_time": product["lead_time"],
                        }
                    )
                elif cumulative_quantity <= d_limit:
                    D += 1
                    product_data["D"].append(
                        {
                            "name": product["name"],
                            "percentage": percent,
                            "lead_time": product["lead_time"],
                        }
                    )
                else:
                    E += 1
                    product_data["E"].append(
                        {
                            "name": product["name"],
                            "percentage": percent,
                            "lead_time": product["lead_time"],
                        }
                    )

                cumulative_quantity += product["lead_time"]

            labels = ["S", "D", "E"]
            output_data = [S, D, E]
            data = {"labels": labels, "data": output_data, "product_data": product_data}

            response = {
                "success": True,
                "message": "Product SDE analysis",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class StockAgingAnalysisView(APIView):
    def get(self, request):
        try:
            view_type = request.GET.get("view_type", "1M")
            if view_type == "1M":
                from_date = datetime.now().date() - timedelta(days=31)
            elif view_type == "3M":
                from_date = datetime.now().date() - timedelta(days=91)
            elif view_type == "6M":
                from_date = datetime.now().date() - timedelta(days=181)
            else:
                from_date = datetime.now().date() - timedelta(days=366)
            total_stock = Stocks.objects.aggregate(total=Sum("quantity"))["total"]
            stocks = Stocks.objects.filter(quantity__gt=0, created_at__lte=from_date)
            old_stock = (
                stocks.aggregate(total=Sum("quantity"))["total"] if stocks else 0
            )
            product_data = (
                stocks.values("product__name")
                .annotate(quantity=(Sum("quantity")))
                .order_by("-quantity")
            )
            ratio_percen = round((old_stock / total_stock) * 100, 2)
            data = {
                "labels": ["Stock Ratio"],
                "data": [ratio_percen],
                "product_data": product_data,
            }
            response = {
                "success": True,
                "message": "Product SDE analysis",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


##report


class ReportMeterial(APIView, CustomPagination):
    serializer_class = ReportMaterialSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(created_at__date__gte=from_date, created_at__date__lte=to_date)
            instance = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values("stock__product__name", "stock__product__ucs_code")
                .annotate(
                    total_stock_in=Coalesce(
                        Sum("quantity", filter=Q(is_stock_out=False)),
                        0,
                        output_field=FloatField(),
                    ),
                    total_stock_out=Coalesce(
                        Sum("quantity", filter=Q(is_stock_out=True)),
                        0,
                        output_field=FloatField(),
                    ),
                    total_stock_in_price=ExpressionWrapper(
                        F("stock__product__price")
                        * Coalesce(Sum("quantity", filter=Q(is_stock_out=False)), 0),
                        output_field=FloatField(),
                    ),
                    total_stock_out_price=ExpressionWrapper(
                        F("stock__product__price")
                        * Coalesce(Sum("quantity", filter=Q(is_stock_out=True)), 0),
                        output_field=FloatField(),
                    ),
                )
                .order_by("stock__product__name")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Material Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportMeterialTransaction(APIView):
    serializer_class = ReportMaterialTransacrionSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(created_at__date__gte=from_date, created_at__date__lte=to_date)
            if product_id := request.GET.get("material_id"):
                product_id_list = product_id.split(",")
                f &= Q(stock__product__id__in=product_id_list)

            instance = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values(
                    "stock__product__name",
                    "created_at__date",
                    "stock__product__ucs_code",
                )
                .annotate(
                    total_stock_in=Coalesce(
                        Sum("quantity", filter=Q(is_stock_out=False)),
                        0,
                        output_field=FloatField(),
                    ),
                    total_stock_out=Coalesce(
                        Sum("quantity", filter=Q(is_stock_out=True)),
                        0,
                        output_field=FloatField(),
                    ),
                )
                .order_by("created_at__date")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Material Transaction Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportFastReceiveMovingItem(APIView):
    serializer_class = ReportFastReceiveMovingItemSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            stock_type = request.GET.get("stock_type", "stockout")

            instance = (
                Products.objects.annotate(
                    quantity=Coalesce(
                        Sum(
                            "stocks__stockshistory__quantity",
                            filter=Q(
                                stocks__stockshistory__is_stock_out=stock_type
                                == "stockout",
                                stocks__stockshistory__created_at__date__gte=from_date,
                                stocks__stockshistory__created_at__date__lte=to_date,
                            ),
                        ),
                        0,
                        output_field=FloatField(),
                    )
                )
                .filter(quantity__gt=0)
                .order_by("-quantity")
                .values("name", "ucs_code", "quantity")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Fast Moving Receive Material Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportCrirticalReorderItem(APIView):
    serializer_class = CrirticalReorderItemSerializer

    def get(self, request):
        try:
            instance = Products.objects.filter(
                net_quantity__lt=F("min_threshold")
            ).order_by("-net_quantity")
            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Critical Reorder Item Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportEmployee(APIView):
    serializer_class = ReportEmployeeSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(
                created_at__date__gte=from_date,
                created_at__date__lte=to_date,
                is_stock_out=True,
                employee__isnull=False,
            )
            if employee_id := request.GET.get("employee_id"):
                employee_id_list = employee_id.split(",")
                f &= Q(employee__id__in=employee_id_list)

            instance = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values(
                    "stock__product__name",
                    "stock__product__ucs_code",
                    "employee__name",
                    "employee__personnel_number",
                )
                .annotate(
                    total_stock_out=Coalesce(
                        Sum("quantity"), 0, output_field=FloatField()
                    ),
                    total_stock_out_price=ExpressionWrapper(
                        F("stock__product__price") * Coalesce(Sum("quantity"), 0),
                        output_field=FloatField(),
                    ),
                )
                .order_by("employee__name")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Employee Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportMaterialUsage(APIView):
    serializer_class = ReportMaterialUsageSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(
                created_at__date__gte=from_date,
                created_at__date__lte=to_date,
                is_stock_out=True,
                purpose__isnull=False,
            )
            if purpose := request.GET.get("purpose"):
                purpose_list = purpose.split(",")
                f &= Q(purpose__in=purpose_list)

            instance = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values(
                    "purpose",
                    "stock__product__name",
                    "created_at__date",
                    "stock__product__ucs_code",
                )
                .annotate(
                    total_stock_out=Coalesce(
                        Sum("quantity"), 0, output_field=FloatField()
                    ),
                )
                .order_by("purpose")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Material Usage Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportDeadStockItem(APIView):
    serializer_class = ReportDeadStockItemSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(
                created_at__date__gte=from_date,
                created_at__date__lte=to_date,
                is_stock_out=True,
            )
            product_ids = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values_list("stock__product__id")
            )
            instance = (
                Products.objects.exclude(id__in=product_ids)
                .order_by("-net_quantity")
                .values("name", "ucs_code", "net_quantity")
            )
            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Critical Reorder Item Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class ReportInventoryException(APIView):
    serializer_class = ReportInventoryExceptionSerializer

    def get(self, request):
        try:
            from_date = request.GET.get("from_date", "")
            if from_date:
                from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            else:
                from_date = datetime.now().date()

            to_date = request.GET.get("to_date", "")
            if to_date:
                to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
            else:
                to_date = from_date - timedelta(days=300)

            f = Q(
                created_at__date__gte=from_date,
                created_at__date__lte=to_date,
                is_stock_out=True,
                obsolete_inventory_barcode__isnull=False,
            )
            if purpose := request.GET.get("purpose"):
                purpose_list = purpose.split(",")
                f &= Q(purpose__in=purpose_list)

            instance = (
                StocksHistory.objects.select_related("stock")
                .filter(f)
                .values(
                    "obsolete_inventory_barcode",
                    "stock__product__name",
                    "created_at__date",
                    "stock__product__ucs_code",
                    "stock__barcode__barcode_no",
                )
                .order_by("-created_at")
            )

            serializer = self.serializer_class(instance, many=True)
            data = serializer.data

            response = {
                "success": True,
                "message": "Material Usage Report List Get Succesfully",
                "data": data,
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)



# forecast

class ProductForecastView(APIView):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Load the trained Prophet models
        self.model_path = Path(settings.BASE_DIR, "MLModel", "prophet_models.pkl")
        self.product_models = joblib.load(self.model_path)

    def get(self, request):
        try:
            # Get forecast type (default: monthly)
            forecast_type = request.query_params.get("type", "monthly").lower()
            if forecast_type not in ["monthly", "quarterly"]:
                return Response(
                    {
                        "success": False,
                        "message": "Invalid type. Use 'monthly' or 'quarterly'.",
                    },
                    status=400,
                )

            # Get number of periods to forecast (default: 3)
            period = int(request.query_params.get("period", 3))
            if forecast_type == "monthly" and period not in [3, 6, 9, 12]:
                return Response(
                    {
                        "success": False,
                        "message": "Invalid period value. Use 3, 6, 9, or 12.",
                    },
                    status=400,
                )

            current_date = datetime.now()

            # Prepare forecast data
            data = []
            for product_name, model in self.product_models.items():
                # Prepare future dates and get forecast
                future_df = self.prepare_future_dates(
                    current_date, period, forecast_type
                )
                forecast_df = model.predict(future_df)
                forecast_data = self.format_forecast_data(
                    forecast_df, period, current_date, forecast_type
                )
                data.append({"product": product_name, "forecast": forecast_data})

            response = {
                "success": True,
                "message": f"{forecast_type.capitalize()} Product Forecast",
                "data": data,
            }
            return Response(response, status=200)

        except Exception as e:
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)

    def prepare_future_dates(self, current_date, period, forecast_type):
        """Prepare future dates based on the forecast type (monthly or quarterly)."""
        if forecast_type == "monthly":
            freq = "MS"  # Month start frequency
            future_dates = pd.date_range(
                start=current_date.replace(day=1),
                periods=period,
                freq=freq,
            )
        else:
            freq = "QS-APR"  # Indian financial year starts in April
            future_dates = pd.date_range(
                start=current_date.replace(day=1),
                periods=period,
                freq=freq,
            )

        future_df = pd.DataFrame({"ds": future_dates})
        return future_df

    def format_forecast_data(self, forecast_df, period, current_date, forecast_type):
        """Format forecast data to include period, actual, lower, upper, start, and end months."""
        forecast_df = forecast_df[
            forecast_df["ds"] >= current_date.replace(day=1)
        ]  # Filter future dates only
        forecast_data = forecast_df.head(period).to_dict("records")
        formatted_data = []

        for row in forecast_data:
            if forecast_type == "monthly":
                period_label = row["ds"].strftime("%b-%Y")  # Format as "Dec-2025"
                formatted_data.append(
                    {
                        "period": period_label,
                        "actual": round(row["yhat"]),
                        "lower": round(row["yhat_lower"]),
                        "upper": round(row["yhat_upper"]),
                    }
                )
            else:
                quarter_start_month, quarter_name = self.get_indian_quarter(row["ds"])
                period_label = f"{row['ds'].year} {quarter_name}"
                end_month = (row["ds"] + pd.DateOffset(months=2)).strftime("%b")
                formatted_data.append(
                    {
                        "period": period_label,
                        "start_month": f"{quarter_start_month}-{row['ds'].year}",
                        "end_month": f"{end_month}-{row['ds'].year}",
                        "actual": round(row["yhat"]),
                        "lower": round(row["yhat_lower"]),
                        "upper": round(row["yhat_upper"]),
                    }
                )

        return formatted_data

    def get_indian_quarter(self, date):
        """Return the Indian financial quarter name and start month."""
        month = date.month
        year = date.year

        if month in [4, 5, 6]:
            return "Apr", "Q1"
        elif month in [7, 8, 9]:
            return "Jul", "Q2"
        elif month in [10, 11, 12]:
            return "Oct", "Q3"
        else:  # January to March
            return "Jan", "Q4"


class TopProductForecastView(APIView):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.model_path = Path(settings.BASE_DIR, "MLModel", "prophet_models.pkl")
        self.product_models = joblib.load(self.model_path)

    def get(self, request):
        try:
            from_date = datetime.now().date()
            to_date = from_date - timedelta(days=30)

            # Get top 10 products by recent stock out quantity
            product_name_list = (
                Products.objects.annotate(
                    quantity=Coalesce(
                        Sum(
                            "stocks__stockshistory__quantity",
                            filter=Q(
                                stocks__stockshistory__is_stock_out=True,
                                stocks__stockshistory__created_at__date__range=(
                                    to_date,
                                    from_date,
                                ),
                            ),
                        ),
                        0,
                    )
                )
                .order_by("-quantity")
                .values_list("name", flat=True)[:10]
            )

            # Prepare forecast for next month
            current_date = datetime.now()
            next_month_start = current_date.replace(day=1) + relativedelta(months=1)
            future_df = pd.DataFrame({"ds": [next_month_start]})

            labels = []
            data = []

            for product_name in product_name_list:
                model = self.product_models.get(product_name)
                if model:
                    forecast_df = model.predict(future_df)
                    yhat = round(forecast_df.iloc[0]["yhat"])
                    labels.append(product_name)
                    data.append(yhat)
                else:
                    labels.append(product_name)
                    data.append(0)  # fallback if model not found

            response = {
                "success": True,
                "message": "Top Product Forecast for Next Month",
                "data": {"labels": labels, "data": data},
            }
            return Response(response, status=200)

        except Exception as e:
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)


class InstantSendDailyReportView(APIView):

    def get(self, request):
        try:
            send_daily_stock_report_mail()
            response = {
                "success": True,
                "message": "Daily Report Mail Sent Successfully",
                "data": [],
            }
            return Response(response, status=200)
        except Exception as e:
            Syserror(e)
            response = {"success": False, "message": str(e)}
            return Response(response, status=400)